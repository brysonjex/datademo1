import argparse
import math
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

UVU_GREEN = "006633"
UVU_DARK_GREEN = "004B2E"
UVU_GRAY = "4D4D4D"
UVU_LIGHT_GRAY = "E6E6E6"
UVU_GOLD = "CBA135"


def leading_digit(value: float) -> int | None:
    if pd.isna(value):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number == 0:
        return None
    number = abs(number)
    while number < 1:
        number *= 10
    digit = int(str(number)[0])
    if digit == 0:
        return None
    return digit


def expected_benford_distribution() -> dict[int, float]:
    return {digit: math.log10(1 + 1 / digit) for digit in range(1, 10)}


def analyze_numeric_column(series: pd.Series, sheet_name: str, column_name: str) -> tuple[pd.DataFrame, dict]:
    digits = series.map(leading_digit).dropna().astype(int)
    total = int(digits.shape[0])
    expected = expected_benford_distribution()

    counts = digits.value_counts().reindex(range(1, 10), fill_value=0).sort_index()
    proportions = counts / total if total > 0 else counts.astype(float)

    detail_rows = []
    for digit in range(1, 10):
        detail_rows.append(
            {
                "sheet": sheet_name,
                "column": column_name,
                "digit": digit,
                "count": int(counts[digit]),
                "proportion": float(proportions[digit]) if total > 0 else 0.0,
                "expected_proportion": expected[digit],
                "difference": float(proportions[digit]) - expected[digit] if total > 0 else -expected[digit],
            }
        )

    if total > 0:
        expected_counts = pd.Series({digit: expected[digit] * total for digit in range(1, 10)})
        chi_square = float(((counts - expected_counts) ** 2 / expected_counts).sum())
        mad = float((proportions - pd.Series(expected)).abs().mean())
    else:
        chi_square = 0.0
        mad = 0.0

    summary = {
        "sheet": sheet_name,
        "column": column_name,
        "total_values": total,
        "chi_square": chi_square,
        "mad": mad,
    }

    return pd.DataFrame(detail_rows), summary


def autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(12, length + 2)


def apply_header_style(cell) -> None:
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor=UVU_GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def main() -> None:
    parser = argparse.ArgumentParser(description="Run Benford's Law analysis and produce an Excel report.")
    parser.add_argument("--input", default="je_samples.xlsx", help="Path to the Excel file.")
    parser.add_argument(
        "--output",
        default="benford_output/benford_report.xlsx",
        help="Path to the Excel output file.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    excel = pd.ExcelFile(input_path)
    detail_frames = []
    summary_rows = []
    all_digits = []

    for sheet_name in excel.sheet_names:
        df = pd.read_excel(input_path, sheet_name=sheet_name)
        numeric_cols = df.select_dtypes(include="number")
        if numeric_cols.empty:
            continue
        for column_name in numeric_cols.columns:
            series = numeric_cols[column_name]
            digits = series.map(leading_digit).dropna().astype(int)
            if not digits.empty:
                all_digits.extend(digits.tolist())
            detail_df, summary = analyze_numeric_column(series, sheet_name, column_name)
            detail_frames.append(detail_df)
            summary_rows.append(summary)

    detail_df = pd.concat(detail_frames, ignore_index=True) if detail_frames else pd.DataFrame(
        columns=["sheet", "column", "digit", "count", "proportion", "expected_proportion", "difference"]
    )
    summary_df = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(
        columns=["sheet", "column", "total_values", "chi_square", "mad"]
    )

    expected = expected_benford_distribution()
    all_digit_series = pd.Series(all_digits) if all_digits else pd.Series(dtype=int)
    overall_counts = all_digit_series.value_counts().reindex(range(1, 10), fill_value=0).sort_index()
    overall_total = int(overall_counts.sum())
    overall_proportions = overall_counts / overall_total if overall_total > 0 else overall_counts.astype(float)

    overall_table = pd.DataFrame(
        {
            "digit": range(1, 10),
            "actual_count": [int(overall_counts[digit]) for digit in range(1, 10)],
            "actual_percent": [float(overall_proportions[digit]) for digit in range(1, 10)],
            "expected_percent": [expected[digit] for digit in range(1, 10)],
        }
    )

    top_deviations = summary_df.sort_values("mad", ascending=False).head(10)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_sheet.merge_cells("A1:G1")
    title_cell = summary_sheet["A1"]
    title_cell.value = "Benford's Law Analysis Report"
    title_cell.font = Font(color="FFFFFF", bold=True, size=16)
    title_cell.fill = PatternFill("solid", fgColor=UVU_DARK_GREEN)
    title_cell.alignment = Alignment(horizontal="center")

    summary_sheet["A2"].value = "Input File"
    summary_sheet["B2"].value = str(input_path)
    summary_sheet["A3"].value = "Generated"
    summary_sheet["B3"].value = datetime.now().strftime("%Y-%m-%d %H:%M")

    for cell in (summary_sheet["A2"], summary_sheet["A3"]):
        cell.font = Font(bold=True, color=UVU_GRAY)

    summary_sheet["A5"].value = "Overall Leading Digit Distribution"
    summary_sheet["A5"].font = Font(bold=True, color=UVU_GREEN, size=12)

    start_row = 6
    headers = ["Digit", "Actual Count", "Actual %", "Expected %"]
    for col_idx, header in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=start_row, column=col_idx, value=header)
        apply_header_style(cell)

    for row_offset, row in enumerate(overall_table.itertuples(index=False), start=1):
        summary_sheet.cell(row=start_row + row_offset, column=1, value=row.digit)
        summary_sheet.cell(row=start_row + row_offset, column=2, value=row.actual_count)
        summary_sheet.cell(row=start_row + row_offset, column=3, value=row.actual_percent)
        summary_sheet.cell(row=start_row + row_offset, column=4, value=row.expected_percent)

    for row in summary_sheet.iter_rows(min_row=start_row + 1, max_row=start_row + 9, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = "0.0%"

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Actual vs Expected Distribution"
    chart.y_axis.title = "Percentage"
    chart.x_axis.title = "Leading Digit"
    data = Reference(summary_sheet, min_col=3, max_col=4, min_row=start_row, max_row=start_row + 9)
    categories = Reference(summary_sheet, min_col=1, min_row=start_row + 1, max_row=start_row + 9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 10
    chart.width = 20
    summary_sheet.add_chart(chart, "F6")

    top_row_start = start_row + 12
    summary_sheet["A" + str(top_row_start)].value = "Top MAD Deviations"
    summary_sheet["A" + str(top_row_start)].font = Font(bold=True, color=UVU_GREEN, size=12)

    top_headers = ["Sheet", "Column", "Total Values", "Chi-Square", "MAD"]
    for col_idx, header in enumerate(top_headers, start=1):
        cell = summary_sheet.cell(row=top_row_start + 1, column=col_idx, value=header)
        apply_header_style(cell)

    for row_offset, row in enumerate(top_deviations.itertuples(index=False), start=1):
        summary_sheet.cell(row=top_row_start + 1 + row_offset, column=1, value=row.sheet)
        summary_sheet.cell(row=top_row_start + 1 + row_offset, column=2, value=row.column)
        summary_sheet.cell(row=top_row_start + 1 + row_offset, column=3, value=int(row.total_values))
        summary_sheet.cell(row=top_row_start + 1 + row_offset, column=4, value=float(row.chi_square))
        summary_sheet.cell(row=top_row_start + 1 + row_offset, column=5, value=float(row.mad))

    mad_chart = BarChart()
    mad_chart.type = "col"
    mad_chart.style = 12
    mad_chart.title = "Top MAD by Column"
    mad_chart.y_axis.title = "MAD"
    mad_chart.x_axis.title = "Column"
    mad_data = Reference(
        summary_sheet,
        min_col=5,
        min_row=top_row_start + 1,
        max_row=top_row_start + 1 + len(top_deviations),
    )
    mad_categories = Reference(
        summary_sheet,
        min_col=2,
        min_row=top_row_start + 2,
        max_row=top_row_start + 1 + len(top_deviations),
    )
    mad_chart.add_data(mad_data, titles_from_data=True)
    mad_chart.set_categories(mad_categories)
    mad_chart.height = 10
    mad_chart.width = 20
    summary_sheet.add_chart(mad_chart, f"F{top_row_start + 1}")

    column_sheet = workbook.create_sheet(title="Column Summary")
    column_headers = ["Sheet", "Column", "Total Values", "Chi-Square", "MAD"]
    for col_idx, header in enumerate(column_headers, start=1):
        cell = column_sheet.cell(row=1, column=col_idx, value=header)
        apply_header_style(cell)

    for row_idx, row in enumerate(summary_df.itertuples(index=False), start=2):
        column_sheet.cell(row=row_idx, column=1, value=row.sheet)
        column_sheet.cell(row=row_idx, column=2, value=row.column)
        column_sheet.cell(row=row_idx, column=3, value=int(row.total_values))
        column_sheet.cell(row=row_idx, column=4, value=float(row.chi_square))
        column_sheet.cell(row=row_idx, column=5, value=float(row.mad))

    detail_sheet = workbook.create_sheet(title="Detail")
    detail_headers = [
        "Sheet",
        "Column",
        "Digit",
        "Count",
        "Proportion",
        "Expected Proportion",
        "Difference",
    ]
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = detail_sheet.cell(row=1, column=col_idx, value=header)
        apply_header_style(cell)

    for row_idx, row in enumerate(detail_df.itertuples(index=False), start=2):
        detail_sheet.cell(row=row_idx, column=1, value=row.sheet)
        detail_sheet.cell(row=row_idx, column=2, value=row.column)
        detail_sheet.cell(row=row_idx, column=3, value=int(row.digit))
        detail_sheet.cell(row=row_idx, column=4, value=int(row.count))
        detail_sheet.cell(row=row_idx, column=5, value=float(row.proportion))
        detail_sheet.cell(row=row_idx, column=6, value=float(row.expected_proportion))
        detail_sheet.cell(row=row_idx, column=7, value=float(row.difference))

    for row in detail_sheet.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = "0.0%"

    for sheet in [summary_sheet, column_sheet, detail_sheet]:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row == 1 and sheet.title == "Summary":
                    continue
                if cell.row == 1:
                    cell.alignment = Alignment(horizontal="center")
        autofit_columns(sheet)

    accent_fill = PatternFill("solid", fgColor=UVU_LIGHT_GRAY)
    for row in summary_sheet.iter_rows(min_row=start_row + 1, max_row=start_row + 9, min_col=1, max_col=4):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = accent_fill

    for row in summary_sheet.iter_rows(
        min_row=top_row_start + 2, max_row=top_row_start + 1 + len(top_deviations), min_col=1, max_col=5
    ):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = accent_fill

    summary_sheet["A5"].fill = PatternFill("solid", fgColor=UVU_GOLD)
    summary_sheet["A5"].font = Font(bold=True, color="FFFFFF")
    summary_sheet["A" + str(top_row_start)].fill = PatternFill("solid", fgColor=UVU_GOLD)
    summary_sheet["A" + str(top_row_start)].font = Font(bold=True, color="FFFFFF")

    workbook.save(output_path)


if __name__ == "__main__":
    main()
